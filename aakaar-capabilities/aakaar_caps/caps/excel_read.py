"""cap.excel_read — read an .xlsx workbook from managed storage into cell grids.

Reads a workbook (``aakaar://`` URI) with openpyxl and returns each sheet's
cells as a 2-D array plus the ordered list of sheet names. Read-only
(``side_effecting=False``). openpyxl is imported lazily so the server can
register the capability without it; the executing host must have the
``automation`` extra installed.
"""

from __future__ import annotations

import io
import logging
from typing import Any

from pydantic import BaseModel, ConfigDict, Field

from aakaar_caps.context import CapabilityContext
from aakaar_caps.spec import CapabilitySpec

logger = logging.getLogger(__name__)
CAP_REF = "cap.excel_read"


class _Inputs(BaseModel):
    model_config = ConfigDict(extra="forbid")
    file_uri: str = Field(description="Managed-storage URI (aakaar://...) of the .xlsx workbook to read.")
    sheet: str | None = Field(
        default=None,
        description="Read only this sheet by name. Omit to read every sheet in the workbook.",
    )
    max_rows: int = Field(
        default=1000, ge=1, le=100000,
        description="Cap on rows returned per sheet (guards against huge workbooks).",
    )


class _Outputs(BaseModel):
    sheets: dict[str, list[list[Any]]] = Field(
        description="Sheet name -> 2-D array of cell values (row-major).",
    )
    sheet_names: list[str] = Field(description="Ordered list of sheet names in the workbook.")


SPEC = CapabilitySpec(
    ref=CAP_REF,
    description=(
        "Read an .xlsx workbook from managed storage with openpyxl and return each sheet's "
        "cells as a 2-D array (row-major) plus the ordered list of sheet names. Optionally "
        "restrict to one sheet and cap rows per sheet. Offline, read-only."
    ),
    input_schema=_Inputs,
    output_schema=_Outputs,
    secrets=(),
    tags=("office", "excel"),
    side_effecting=False,
)


def _require_openpyxl() -> Any:
    try:
        import openpyxl  # type: ignore[import-untyped]
    except ImportError as exc:  # pragma: no cover - exercised only when dep absent
        raise RuntimeError(
            "cap.excel_read needs openpyxl — install aakaar-capabilities[automation]"
        ) from exc
    return openpyxl


async def run(ctx: CapabilityContext, inputs: dict[str, Any]) -> dict[str, Any]:
    openpyxl = _require_openpyxl()
    file_uri = inputs["file_uri"]
    only = inputs.get("sheet")
    max_rows = int(inputs.get("max_rows", 1000))

    data = await ctx.read_object(file_uri)
    logger.info("cap.excel_read start run_id=%s uri=%s sheet=%s", ctx.run_id, file_uri, only)
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        names = list(wb.sheetnames)
        targets = [only] if only else names
        if only and only not in names:
            raise ValueError(f"sheet {only!r} not found; available: {names}")

        sheets: dict[str, list[list[Any]]] = {}
        for name in targets:
            ws = wb[name]
            rows: list[list[Any]] = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= max_rows:
                    break
                rows.append(list(row))
            sheets[name] = rows
    finally:
        wb.close()

    logger.info("cap.excel_read ok run_id=%s sheets=%d", ctx.run_id, len(sheets))
    return {"sheets": sheets, "sheet_names": names}
