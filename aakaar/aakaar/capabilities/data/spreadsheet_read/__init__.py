"""cap.spreadsheet_read — read an xlsx/csv spreadsheet into structured rows.

Server-local, read-only extraction over a single object in the tenant's
object store (an ``aakaar://`` URI the upstream graph already produced —
typically from ``cap.file_download`` or an email attachment). Unlike
``cap.doc_extract`` (which dispatches across many document types) and
``cap.data_transform`` (which runs a pandas op pipeline), this capability is
a focused, pandas-free spreadsheet reader: it streams the sheet with a strict
row cap so a large workbook can be ingested for a banking reconciliation
without materialising the whole frame in memory.

  - csv / tsv -> stdlib ``csv`` (no third-party dependency at all).
  - xlsx / xlsm -> ``openpyxl`` in read-only/streaming mode, behind the
    optional ``office`` extra. The import is LAZY so this module imports fine
    on a host without openpyxl; the handler raises a clear RuntimeError naming
    the extra to install (``pip install 'aakaar[office]'``).

The first row is treated as the header unless ``has_header`` is false, in
which case columns are named ``col_1``, ``col_2``, …. Output is a list of row
dicts plus the ordered column names and a ``truncated`` flag set when the
sheet had more rows than ``max_rows`` allowed.

Caps (defence against a memory/CPU bomb):
  - the encoded source size is bounded before any parsing;
  - at most ``max_rows`` data rows are returned (hard-capped by ``_ROW_CAP``);
  - at most ``_MAX_COLS`` columns are read.

Read-only: no secrets, no network, ``side_effecting=False``.
"""

from __future__ import annotations

import csv
import io
import logging
from typing import Any

from pydantic import BaseModel, ConfigDict, Field

from aakaar.interpreter.activities.types import ActivityContext
from aakaar.shared.registry import CapabilityDefinition
from aakaar.storage.object_store import parse_uri

logger = logging.getLogger(__name__)
CAP_REF = "cap.spreadsheet_read"

# A spreadsheet is read fully into memory before streaming, so an unbounded
# source is a memory bomb (xlsx is a zip — a small file can expand hugely);
# refuse early on the encoded size.
_MAX_SOURCE_BYTES = 32 * 1024 * 1024  # 32 MiB on the wire
# Absolute ceiling on returned rows regardless of the caller's `max_rows`, so a
# hand-written DAG can't ask for an unbounded result set.
_ROW_CAP = 1_000_000
_DEFAULT_MAX_ROWS = 100_000
# A wildly wide sheet (e.g. a malformed file reporting millions of columns) is
# refused before we allocate per-column structures.
_MAX_COLS = 2_048

_EXT_KINDS = {
    "csv": "csv",
    "tsv": "csv",
    "xlsx": "xlsx",
    "xlsm": "xlsx",
}


class _Inputs(BaseModel):
    model_config = ConfigDict(extra="forbid")
    source: str = Field(
        description="aakaar:// URI of the spreadsheet (csv/tsv/xlsx/xlsm) to read.",
    )
    source_format: str | None = Field(
        default=None,
        description=(
            "Override the format: one of 'csv', 'xlsx'. When null the source "
            "URI's file extension decides (.csv/.tsv -> csv, .xlsx/.xlsm -> xlsx)."
        ),
    )
    sheet: str | None = Field(
        default=None,
        description=(
            "For xlsx: the worksheet name to read. Null reads the first "
            "(active-order) sheet. Ignored for csv."
        ),
    )
    has_header: bool = Field(
        default=True,
        description=(
            "When true the first row supplies the column names. When false "
            "columns are auto-named col_1, col_2, … and every row is data."
        ),
    )
    max_rows: int = Field(
        default=_DEFAULT_MAX_ROWS,
        ge=1,
        le=_ROW_CAP,
        description=(
            "Maximum number of data rows to return. Extra rows are dropped and "
            "`truncated` is set. Hard-capped server-side."
        ),
    )
    encoding: str = Field(
        default="utf-8",
        description="Text decoding for csv/tsv content. Decoding is error-tolerant.",
    )


class _Outputs(BaseModel):
    columns: list[str] = Field(description="Ordered column names.")
    rows: list[dict[str, Any]] = Field(
        description="Data rows as {column: value} dicts, in sheet order.",
    )
    row_count: int = Field(description="Number of rows returned (after any cap).")
    sheet: str | None = Field(
        default=None,
        description="Name of the worksheet read (xlsx only; null for csv).",
    )
    truncated: bool = Field(
        description="True when the sheet had more rows than `max_rows` allowed.",
    )


definition = CapabilityDefinition(
    ref=CAP_REF,
    description=(
        "Read an xlsx/xlsm/csv/tsv spreadsheet from the object store into "
        "structured rows (list of {column: value} dicts) with the ordered "
        "column names and a truncation flag. Streams with a strict row cap; "
        "csv uses the stdlib, xlsx uses openpyxl (the 'office' extra, "
        "lazy-imported). Read-only — no secrets, no network."
    ),
    input_schema=_Inputs,
    output_schema=_Outputs,
    # Pure read-only extraction: nothing escapes the run sandbox. Runs for real
    # even in a dry-run so a simulated plan still yields the real rows.
    side_effecting=False,
    secrets=(),
    tags=("data", "spreadsheet", "xlsx", "csv", "tabular", "read"),
)


# ---------------------------------------------------------------------------
# Pure helpers (no ctx / no I/O) — unit-testable in isolation.
# ---------------------------------------------------------------------------


def _detect_kind(uri: str, override: str | None) -> str:
    if override:
        kind = override.strip().lower()
        if kind not in {"csv", "xlsx"}:
            raise ValueError(
                f"cap.spreadsheet_read: unsupported source_format {override!r}; "
                "expected 'csv' or 'xlsx'"
            )
        return kind
    _, key = parse_uri(uri)
    ext = key.rsplit(".", 1)[-1].lower() if "." in key else ""
    inferred = _EXT_KINDS.get(ext)
    if inferred is None:
        raise ValueError(
            f"cap.spreadsheet_read: cannot infer format from {uri!r}; set "
            "`source_format` explicitly ('csv' or 'xlsx')"
        )
    return inferred


def _header_for(width: int, first_row: list[Any] | None, has_header: bool) -> list[str]:
    """Build the column names for a sheet `width` columns wide.

    With a header row, blank/missing header cells get a positional fallback
    (``col_N``) and duplicate names are disambiguated with a ``_N`` suffix so
    the row dicts never silently drop a column.
    """
    if has_header and first_row is not None:
        names: list[str] = []
        seen: dict[str, int] = {}
        for i in range(width):
            raw = first_row[i] if i < len(first_row) else None
            name = "" if raw is None else str(raw).strip()
            if not name:
                name = f"col_{i + 1}"
            if name in seen:
                seen[name] += 1
                name = f"{name}_{seen[name]}"
            else:
                seen[name] = 0
            names.append(name)
        return names
    return [f"col_{i + 1}" for i in range(width)]


def _rows_from_matrix(
    matrix: list[list[Any]], *, has_header: bool, max_rows: int
) -> tuple[list[str], list[dict[str, Any]], bool]:
    """Turn a list-of-lists sheet into (columns, row dicts, truncated).

    Pure: the caller is responsible for streaming/capping the source. `matrix`
    is the already-bounded set of raw rows (header included when present).
    """
    if not matrix:
        return [], [], False
    width = max(len(r) for r in matrix)
    if width > _MAX_COLS:
        raise RuntimeError(
            f"cap.spreadsheet_read: sheet has {width} columns, exceeding the "
            f"{_MAX_COLS}-column limit"
        )
    header_source = matrix[0] if has_header else None
    columns = _header_for(width, header_source, has_header)
    data_rows = matrix[1:] if has_header else matrix
    truncated = len(data_rows) > max_rows
    out: list[dict[str, Any]] = []
    for raw in data_rows[:max_rows]:
        row = {columns[i]: (raw[i] if i < len(raw) else None) for i in range(width)}
        out.append(row)
    return columns, out, truncated


def _read_csv_matrix(
    raw: bytes, *, encoding: str, delimiter: str, row_budget: int
) -> tuple[list[list[Any]], bool]:
    """Decode + parse a CSV/TSV into a bounded list-of-lists.

    Reads at most `row_budget` rows (header included) plus one extra row to
    detect truncation, so we never materialise the full file when it overflows.
    """
    text = raw.decode(encoding or "utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    matrix: list[list[Any]] = []
    overflow = False
    for row in reader:
        if len(matrix) >= row_budget:
            overflow = True
            break
        matrix.append(list(row))
    return matrix, overflow


def _read_xlsx_matrix(
    raw: bytes, *, sheet: str | None, row_budget: int
) -> tuple[list[list[Any]], bool, str]:
    """Stream an xlsx worksheet into a bounded list-of-lists.

    openpyxl is imported lazily so this module loads without the 'office' extra.
    The workbook is opened read-only so rows stream rather than fully load.
    Returns (matrix, overflow, sheet_name).
    """
    try:
        import openpyxl  # noqa: PLC0415  (lazy: optional 'office' extra)
    except ImportError as e:
        raise RuntimeError(
            "cap.spreadsheet_read: reading xlsx requires the 'openpyxl' package, "
            "which is not installed on this worker host. Install it with "
            "`pip install 'aakaar[office]'` (or `pip install openpyxl`)."
        ) from e

    wb = openpyxl.load_workbook(
        io.BytesIO(raw), read_only=True, data_only=True
    )
    try:
        if sheet is not None:
            if sheet not in wb.sheetnames:
                raise RuntimeError(
                    f"cap.spreadsheet_read: workbook has no sheet named {sheet!r}; "
                    f"available: {wb.sheetnames!r}"
                )
            ws = wb[sheet]
        else:
            ws = wb.active
            if ws is None:
                raise RuntimeError(
                    "cap.spreadsheet_read: workbook has no active worksheet; "
                    "pass `sheet` to name one explicitly"
                )
        sheet_name = str(ws.title)
        matrix: list[list[Any]] = []
        overflow = False
        for row in ws.iter_rows(values_only=True):
            if len(matrix) >= row_budget:
                overflow = True
                break
            matrix.append(list(row))
        return matrix, overflow, sheet_name
    finally:
        # read-only workbooks hold an open zip handle; close it deterministically.
        wb.close()


# ---------------------------------------------------------------------------
# Handler
# ---------------------------------------------------------------------------


async def handler(ctx: ActivityContext, inputs: dict[str, Any]) -> dict[str, Any]:
    source = inputs["source"]
    kind = _detect_kind(source, inputs.get("source_format"))
    has_header = bool(inputs.get("has_header", True))
    max_rows = int(inputs.get("max_rows", _DEFAULT_MAX_ROWS))
    max_rows = min(max_rows, _ROW_CAP)
    encoding = inputs.get("encoding") or "utf-8"
    sheet = inputs.get("sheet")

    logger.info(
        "cap.spreadsheet_read start run_id=%s source=%s kind=%s max_rows=%d",
        ctx.run_id,
        source,
        kind,
        max_rows,
    )

    raw = ctx.object_store.get(source)
    if len(raw) > _MAX_SOURCE_BYTES:
        raise RuntimeError(
            f"cap.spreadsheet_read: source is {len(raw)} bytes, exceeding the "
            f"{_MAX_SOURCE_BYTES}-byte limit"
        )

    # Budget = max data rows (+1 for a header when present, +1 to detect overflow
    # without reading the whole file).
    row_budget = max_rows + (1 if has_header else 0) + 1
    sheet_name: str | None = None
    if kind == "csv":
        _, key = parse_uri(source)
        delimiter = "\t" if key.lower().endswith(".tsv") else ","
        matrix, overflow = _read_csv_matrix(
            raw, encoding=encoding, delimiter=delimiter, row_budget=row_budget
        )
    else:
        matrix, overflow, sheet_name = _read_xlsx_matrix(
            raw, sheet=sheet, row_budget=row_budget
        )

    columns, rows, capped = _rows_from_matrix(
        matrix, has_header=has_header, max_rows=max_rows
    )
    truncated = overflow or capped

    logger.info(
        "cap.spreadsheet_read ok run_id=%s source=%s rows=%d cols=%d truncated=%s",
        ctx.run_id,
        source,
        len(rows),
        len(columns),
        truncated,
    )
    return {
        "columns": columns,
        "rows": rows,
        "row_count": len(rows),
        "sheet": sheet_name,
        "truncated": truncated,
    }
