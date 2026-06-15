"""Tests for cap.spreadsheet_read.

CSV reading is pure-stdlib and always exercised. The xlsx happy path needs
``openpyxl`` (the 'office' extra); it is installed in this repo's core deps so
the test runs, but the missing-dependency degradation is asserted by forcing
the lazy import to fail. Definition shape, format detection, the row cap, and
the header/uniqueness helpers are always exercised.
"""

from __future__ import annotations

import io
import uuid
from pathlib import Path

import pytest
from pydantic import ValidationError

from aakaar.capabilities.data.spreadsheet_read import (
    CAP_REF,
    _detect_kind,
    _header_for,
    _read_xlsx_matrix,
    _rows_from_matrix,
    definition,
    handler,
)
from aakaar.interpreter.activities.types import ActivityContext


def _ctx(tmp_path: Path) -> ActivityContext:
    from aakaar.shared.registry import build_default_registry
    from aakaar.storage import LocalFsObjectStore
    from aakaar.vault import LocalVault

    return ActivityContext(
        tenant_id=uuid.uuid4(),
        run_id=uuid.uuid4(),
        registry=build_default_registry(),
        object_store=LocalFsObjectStore(tmp_path / "objs"),
        vault=LocalVault(tmp_path / "vault"),
    )


def _xlsx_bytes(rows: list[list[object]], *, title: str = "Sheet1") -> bytes:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------------
# Definition + input validation
# --------------------------------------------------------------------------


def test_definition_shape() -> None:
    assert definition.ref == CAP_REF == "cap.spreadsheet_read"
    assert definition.secrets == ()
    assert "spreadsheet" in definition.tags
    assert set(definition.output_schema.model_fields) == {
        "columns",
        "rows",
        "row_count",
        "sheet",
        "truncated",
    }


def test_definition_is_read_only() -> None:
    assert definition.side_effecting is False


def test_input_schema_forbids_extra_keys() -> None:
    with pytest.raises(ValidationError):
        definition.input_schema(source="aakaar://t/x/y.csv", bogus=1)


def test_input_schema_requires_source() -> None:
    with pytest.raises(ValidationError):
        definition.input_schema()


def test_input_schema_caps_max_rows() -> None:
    with pytest.raises(ValidationError):
        definition.input_schema(source="aakaar://t/x/y.csv", max_rows=10_000_000)


# --------------------------------------------------------------------------
# Pure helpers
# --------------------------------------------------------------------------


def test_detect_kind_from_extension() -> None:
    assert _detect_kind("aakaar://t/x/a.csv", None) == "csv"
    assert _detect_kind("aakaar://t/x/a.tsv", None) == "csv"
    assert _detect_kind("aakaar://t/x/a.xlsx", None) == "xlsx"
    assert _detect_kind("aakaar://t/x/a.xlsm", None) == "xlsx"


def test_detect_kind_override_wins() -> None:
    assert _detect_kind("aakaar://t/x/a.bin", "csv") == "csv"


def test_detect_kind_unknown_raises() -> None:
    with pytest.raises(ValueError, match="cannot infer format"):
        _detect_kind("aakaar://t/x/a.bin", None)


def test_header_dedupes_and_fills_blanks() -> None:
    cols = _header_for(4, ["id", "", "id", None], has_header=True)
    # blanks -> positional fallback; duplicate "id" -> "id_1".
    assert cols == ["id", "col_2", "id_1", "col_4"]


def test_header_autonames_without_header() -> None:
    assert _header_for(3, None, has_header=False) == ["col_1", "col_2", "col_3"]


def test_rows_from_matrix_pads_short_rows() -> None:
    matrix = [["a", "b", "c"], [1, 2], [4, 5, 6]]
    cols, rows, trunc = _rows_from_matrix(matrix, has_header=True, max_rows=10)
    assert cols == ["a", "b", "c"]
    assert rows == [{"a": 1, "b": 2, "c": None}, {"a": 4, "b": 5, "c": 6}]
    assert trunc is False


def test_rows_from_matrix_truncates() -> None:
    matrix = [["a"], [1], [2], [3]]
    cols, rows, trunc = _rows_from_matrix(matrix, has_header=True, max_rows=2)
    assert rows == [{"a": 1}, {"a": 2}]
    assert trunc is True


# --------------------------------------------------------------------------
# CSV happy path (pure stdlib, always runs)
# --------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_csv_happy_path(tmp_path: Path) -> None:
    ctx = _ctx(tmp_path)
    data = b"txn_id,amount,status\n1,100.50,OK\n2,200,FAIL\n"
    stored = ctx.object_store.put(str(ctx.tenant_id), "recon/in.csv", data)

    out = await handler(ctx, {"source": stored.uri})
    assert out["columns"] == ["txn_id", "amount", "status"]
    assert out["rows"] == [
        {"txn_id": "1", "amount": "100.50", "status": "OK"},
        {"txn_id": "2", "amount": "200", "status": "FAIL"},
    ]
    assert out["row_count"] == 2
    assert out["truncated"] is False
    assert out["sheet"] is None


@pytest.mark.asyncio
async def test_tsv_uses_tab_delimiter(tmp_path: Path) -> None:
    ctx = _ctx(tmp_path)
    data = b"a\tb\n1\t2\n"
    stored = ctx.object_store.put(str(ctx.tenant_id), "recon/in.tsv", data)
    out = await handler(ctx, {"source": stored.uri})
    assert out["columns"] == ["a", "b"]
    assert out["rows"] == [{"a": "1", "b": "2"}]


@pytest.mark.asyncio
async def test_csv_truncates_to_max_rows(tmp_path: Path) -> None:
    ctx = _ctx(tmp_path)
    data = b"a\n" + b"\n".join(str(i).encode() for i in range(10)) + b"\n"
    stored = ctx.object_store.put(str(ctx.tenant_id), "recon/big.csv", data)
    out = await handler(ctx, {"source": stored.uri, "max_rows": 3})
    assert out["row_count"] == 3
    assert out["truncated"] is True


# --------------------------------------------------------------------------
# xlsx happy path + sheet selection (needs openpyxl; installed in core)
# --------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_xlsx_happy_path(tmp_path: Path) -> None:
    pytest.importorskip("openpyxl")
    ctx = _ctx(tmp_path)
    data = _xlsx_bytes(
        [["txn_id", "amount"], [1, 100.5], [2, 200]], title="Recon"
    )
    stored = ctx.object_store.put(str(ctx.tenant_id), "recon/in.xlsx", data)

    out = await handler(ctx, {"source": stored.uri})
    assert out["columns"] == ["txn_id", "amount"]
    assert out["rows"] == [
        {"txn_id": 1, "amount": 100.5},
        {"txn_id": 2, "amount": 200},
    ]
    assert out["sheet"] == "Recon"
    assert out["truncated"] is False


@pytest.mark.asyncio
async def test_xlsx_named_sheet_and_missing_sheet(tmp_path: Path) -> None:
    pytest.importorskip("openpyxl")
    import openpyxl

    wb = openpyxl.Workbook()
    wb.active.title = "First"
    wb.active.append(["a"])
    wb.active.append([1])
    ws2 = wb.create_sheet("Second")
    ws2.append(["b"])
    ws2.append([2])
    buf = io.BytesIO()
    wb.save(buf)

    ctx = _ctx(tmp_path)
    stored = ctx.object_store.put(str(ctx.tenant_id), "wb.xlsx", buf.getvalue())

    out = await handler(ctx, {"source": stored.uri, "sheet": "Second"})
    assert out["sheet"] == "Second"
    assert out["columns"] == ["b"]
    assert out["rows"] == [{"b": 2}]

    with pytest.raises(RuntimeError, match="no sheet named"):
        await handler(ctx, {"source": stored.uri, "sheet": "Nope"})


# --------------------------------------------------------------------------
# Safety property: oversized source is refused before parsing
# --------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_rejects_oversized_source(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    import aakaar.capabilities.data.spreadsheet_read as mod

    monkeypatch.setattr(mod, "_MAX_SOURCE_BYTES", 16)
    ctx = _ctx(tmp_path)
    stored = ctx.object_store.put(str(ctx.tenant_id), "big.csv", b"x" * 64)
    with pytest.raises(RuntimeError, match="exceeding the"):
        await handler(ctx, {"source": stored.uri})


def test_rows_from_matrix_rejects_too_many_columns(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import aakaar.capabilities.data.spreadsheet_read as mod

    monkeypatch.setattr(mod, "_MAX_COLS", 2)
    with pytest.raises(RuntimeError, match="exceeding the"):
        _rows_from_matrix([["a", "b", "c"], [1, 2, 3]], has_header=True, max_rows=10)


# --------------------------------------------------------------------------
# Missing-optional-dependency degradation (openpyxl absent)
# --------------------------------------------------------------------------


def test_xlsx_missing_openpyxl_raises_actionable(monkeypatch: pytest.MonkeyPatch) -> None:
    """Module imports fine without openpyxl; the xlsx path raises a clear,
    actionable RuntimeError naming the 'office' extra."""
    import builtins

    real_import = builtins.__import__

    def _fake_import(name: str, *args: object, **kwargs: object) -> object:
        if name == "openpyxl":
            raise ImportError("no module named openpyxl")
        return real_import(name, *args, **kwargs)

    monkeypatch.setattr(builtins, "__import__", _fake_import)
    with pytest.raises(RuntimeError, match="office"):
        _read_xlsx_matrix(b"PK\x03\x04not-a-real-xlsx", sheet=None, row_budget=10)
